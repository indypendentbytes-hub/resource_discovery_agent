"""Read, normalize, and validate the governed workbook."""

from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from resource_discovery_agent.models import (
    AudienceClassification,
    CostClassification,
    CostModel,
    GoalType,
    InclusionBasis,
    ResourceType,
    SourceRecord,
    SourceStatus,
    UserProfile,
    UserType,
)

REQUIRED_SHEETS = (
    "Overview",
    "Source Catalog",
    "Resource Categories",
    "Role Resource Map",
    "Agent Routing",
    "Developer Integration",
    "Ingestion Roadmap",
)
REQUIRED_SOURCE_COLUMNS = (
    "Source ID",
    "Source Name",
    "Canonical URL",
    "Organization",
    "Jurisdiction",
    "Authority Class",
    "Source Format",
    "Primary Resource Category",
    "Capability Tags",
    "Best-Fit User Roles",
    "Core User Value",
    "Refresh Cadence",
    "Citation Rule",
    "Status",
)


@dataclass(frozen=True, slots=True)
class ValidationIssue:
    code: str
    message: str
    severity: str = "error"
    sheet: str | None = None
    row: int | None = None


@dataclass(slots=True)
class CatalogValidationReport:
    workbook: Path
    source_count: int
    guide_reported_count: int | None
    issues: list[ValidationIssue] = field(default_factory=list)

    @property
    def errors(self) -> list[ValidationIssue]:
        return [issue for issue in self.issues if issue.severity == "error"]

    @property
    def warnings(self) -> list[ValidationIssue]:
        return [issue for issue in self.issues if issue.severity == "warning"]

    @property
    def is_valid(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        return {
            "workbook": str(self.workbook),
            "source_count": self.source_count,
            "guide_reported_count": self.guide_reported_count,
            "is_valid": self.is_valid,
            "errors": [asdict(issue) for issue in self.errors],
            "warnings": [asdict(issue) for issue in self.warnings],
        }


def _iter_records(sheet: Worksheet) -> Iterable[tuple[int, dict[str, Any]]]:
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    for row_number, values in enumerate(rows, start=2):
        record = {key: value for key, value in zip(headers, values, strict=False) if key}
        if any(value not in (None, "") for value in record.values()):
            yield row_number, record


def _split(value: Any) -> list[str]:
    return [part.strip() for part in str(value or "").split(";") if part.strip()]


def normalize_status(value: Any) -> SourceStatus:
    status = str(value or "").strip().casefold()
    if status == "active":
        return SourceStatus.ACTIVE
    if "closed" in status:
        return SourceStatus.CLOSED
    if any(term in status for term in ("histor", "archiv", "legacy", "stale", "supersed")):
        return SourceStatus.ARCHIVED
    if any(term in status for term in ("quarant", "needs review", "malformed", "unresolved")):
        return SourceStatus.QUARANTINED
    if any(term in status for term in ("verify", "verification", "manual", "contradictory")):
        return SourceStatus.VERIFICATION_REQUIRED
    if "active" in status or "current" in status:
        return SourceStatus.ACTIVE
    return SourceStatus.UNVERIFIED


def _resource_types(row: dict[str, Any]) -> list[ResourceType]:
    text = " ".join(
        str(row.get(key) or "")
        for key in ("Primary Resource Category", "Source Format", "Authority Class")
    ).casefold()
    types: set[ResourceType] = set()
    mappings = (
        (("dataset", "statistics", "data portal", "gis"), ResourceType.PUBLIC_DATASET),
        (("funding", "grant", "loan", "financial"), ResourceType.FUNDING_OPPORTUNITY),
        (("training", "education", "course"), ResourceType.TRAINING_PROGRAM),
        (("technical assistance",), ResourceType.TECHNICAL_ASSISTANCE_PROGRAM),
        (("mentor",), ResourceType.MENTORSHIP_PROGRAM),
        (("accelerator", "incubator"), ResourceType.ACCELERATOR_OR_INCUBATOR),
        (("procurement",), ResourceType.PROCUREMENT_OPPORTUNITY),
        (("equipment", "facility", "infrastructure"), ResourceType.FACILITY_OR_EQUIPMENT_RESOURCE),
        (("regulation", "compliance", "legal"), ResourceType.LEGAL_OR_REGULATORY_RESOURCE),
        (("research", "market intelligence"), ResourceType.RESEARCH_OR_MARKET_INTELLIGENCE),
        (("commercial",), ResourceType.COMMERCIAL_TOOL),
        (("university", "extension"), ResourceType.UNIVERSITY_PROGRAM),
        (("government", "official"), ResourceType.GOVERNMENT_PROGRAM),
        (("nonprofit",), ResourceType.NONPROFIT_PROGRAM),
    )
    for terms, resource_type in mappings:
        if any(term in text for term in terms):
            types.add(resource_type)
    return sorted(types or {ResourceType.FREE_PUBLIC_RESOURCE}, key=lambda item: item.value)


def _inclusion_basis(types: list[ResourceType]) -> list[InclusionBasis]:
    bases: set[InclusionBasis] = set()
    if ResourceType.FUNDING_OPPORTUNITY in types:
        bases.add(InclusionBasis.FUNDING_OR_FINANCIAL_ACCESS)
    if any(
        item in types
        for item in (
            ResourceType.PUBLIC_DATASET,
            ResourceType.RESEARCH_OR_MARKET_INTELLIGENCE,
            ResourceType.LEGAL_OR_REGULATORY_RESOURCE,
        )
    ):
        bases.add(InclusionBasis.USEFUL_PUBLIC_INFORMATION)
    if ResourceType.FACILITY_OR_EQUIPMENT_RESOURCE in types:
        bases.add(InclusionBasis.BUSINESS_ENABLING_INFRASTRUCTURE)
    if any(
        item in types
        for item in (
            ResourceType.ACCELERATOR_OR_INCUBATOR,
            ResourceType.PROCUREMENT_OPPORTUNITY,
        )
    ):
        bases.add(InclusionBasis.OPPORTUNITY_ACCESS)
    if not bases:
        bases.add(InclusionBasis.DIRECT_ASSISTANCE)
    return sorted(bases, key=lambda item: item.value)


def _audiences(row: dict[str, Any]) -> AudienceClassification:
    text = str(row.get("Best-Fit User Roles") or "").casefold()
    mappings = {
        "grower": UserType.GROWER,
        "business": UserType.BUSINESS_OWNER,
        "nonprofit": UserType.NONPROFIT,
        "worker": UserType.WORKER,
        "student": UserType.STUDENT,
        "research": UserType.RESEARCHER,
        "investor": UserType.INVESTOR,
        "community": UserType.COMMUNITY_ORGANIZATION,
        "employer": UserType.EMPLOYER,
        "restaurant": UserType.FOOD_BUSINESS,
        "institution": UserType.FOOD_BUSINESS,
    }
    users = sorted(
        {user_type for term, user_type in mappings.items() if term in text},
        key=lambda item: item.value,
    )
    category = str(row.get("Primary Resource Category") or "").casefold()
    goals: set[GoalType] = set()
    if "fund" in category or "capital" in category:
        goals.add(GoalType.FIND_FUNDING)
    if "data" in category or "statistics" in category or "intelligence" in category:
        goals.add(GoalType.ACCESS_DATA)
    if "training" in category or "certification" in category:
        goals.add(GoalType.OBTAIN_CERTIFICATION)
    if "property" in category or "land" in category:
        goals.add(GoalType.ACQUIRE_PROPERTY)
    return AudienceClassification(
        user_types=users,
        goals=sorted(goals, key=lambda item: item.value),
    )


class WorkbookCatalog:
    def __init__(
        self,
        workbook_path: Path,
        additions_dir: Path | None = None,
    ) -> None:
        self.workbook_path = workbook_path
        self.additions_dir = additions_dir or workbook_path.parent / "curated-additions"
        self._sources: dict[str, SourceRecord] | None = None

    def _load(self) -> dict[str, SourceRecord]:
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            sources: dict[str, SourceRecord] = {}
            for _, row in _iter_records(workbook["Source Catalog"]):
                source_id = str(row.get("Source ID") or "").strip()
                if not source_id:
                    continue
                status = normalize_status(row.get("Status"))
                warning = str(row.get("Restrictions / Warnings") or "").strip()
                reviewed = row.get("Last Reviewed")
                types = _resource_types(row)
                sources[source_id] = SourceRecord(
                    source_id=source_id,
                    official_name=str(row.get("Source Name") or "").strip(),
                    submitted_url=row.get("Submitted URL") or None,
                    canonical_url=str(row.get("Canonical URL") or "unresolved"),
                    organization=str(row.get("Organization") or "").strip(),
                    jurisdiction=str(row.get("Jurisdiction") or "").strip(),
                    service_geographies=_split(row.get("Jurisdiction")),
                    authority_class=str(row.get("Authority Class") or "").strip(),
                    source_format=str(row.get("Source Format") or "").strip(),
                    resource_category=str(row.get("Primary Resource Category") or "").strip(),
                    resource_types=types,
                    inclusion_basis=_inclusion_basis(types),
                    audiences=_audiences(row),
                    cost=CostClassification(model=CostModel.UNKNOWN),
                    capability_tags=_split(row.get("Capability Tags")),
                    core_user_value=str(row.get("Core User Value") or "").strip(),
                    exclusions=[warning] if warning else [],
                    status=status,
                    last_verified=reviewed.date() if isinstance(reviewed, datetime) else None,
                    refresh_trigger=str(row.get("Refresh Cadence") or "").strip(),
                    citation_rule=str(row.get("Citation Rule") or "").strip(),
                    known_failure_modes=[warning] if warning else [],
                    live_verification_required=(
                        status is not SourceStatus.ACTIVE
                        or "live" in str(row.get("Integration Pattern") or "").casefold()
                    ),
                )
        finally:
            workbook.close()
        if self.additions_dir.exists():
            for path in sorted(self.additions_dir.glob("*.json")):
                addition = SourceRecord.model_validate_json(path.read_text(encoding="utf-8"))
                if addition.source_id in sources:
                    raise ValueError(f"Curated addition duplicates source ID {addition.source_id}")
                sources[addition.source_id] = addition
        return sources

    @property
    def sources(self) -> dict[str, SourceRecord]:
        if self._sources is None:
            self._sources = self._load()
        return self._sources

    def get(self, source_id: str) -> SourceRecord | None:
        return self.sources.get(source_id)

    def find_candidates(
        self, *, profile: UserProfile, goal: GoalType, limit: int = 20
    ) -> list[SourceRecord]:
        ranked = sorted(
            self.sources.values(),
            key=lambda source: (
                goal not in source.audiences.goals,
                profile.user_type not in source.audiences.user_types,
                source.source_id,
            ),
        )
        return ranked[:limit]


def validate_workbook(
    workbook_path: Path, *, guide_reported_count: int | None = 191
) -> CatalogValidationReport:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    issues: list[ValidationIssue] = []
    source_count = 0
    try:
        for name in REQUIRED_SHEETS:
            if name not in workbook.sheetnames:
                issues.append(
                    ValidationIssue(
                        "missing_sheet",
                        f"Missing required sheet: {name}",
                        sheet=name,
                    )
                )
        if "Source Catalog" not in workbook.sheetnames:
            return CatalogValidationReport(workbook_path, 0, guide_reported_count, issues)
        sheet = workbook["Source Catalog"]
        headers = {str(cell.value).strip() for cell in sheet[1] if cell.value is not None}
        for column in REQUIRED_SOURCE_COLUMNS:
            if column not in headers:
                issues.append(
                    ValidationIssue(
                        "missing_column",
                        f"Missing Source Catalog column: {column}",
                        sheet="Source Catalog",
                    )
                )
        categories = {
            str(row.get("Resource Category")).strip()
            for _, row in _iter_records(workbook["Resource Categories"])
            if row.get("Resource Category")
        }
        seen: set[str] = set()
        for row_number, row in _iter_records(sheet):
            source_id = str(row.get("Source ID") or "").strip()
            if not source_id:
                continue
            source_count += 1
            if not re.fullmatch(r"SRC-\d{3,}", source_id):
                issues.append(
                    ValidationIssue(
                        "invalid_source_id",
                        source_id,
                        sheet="Source Catalog",
                        row=row_number,
                    )
                )
            if source_id in seen:
                issues.append(
                    ValidationIssue(
                        "duplicate_source_id",
                        source_id,
                        sheet="Source Catalog",
                        row=row_number,
                    )
                )
            seen.add(source_id)
            parsed = urlparse(str(row.get("Canonical URL") or ""))
            if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                issues.append(
                    ValidationIssue(
                        "invalid_canonical_url",
                        f"{source_id} has no valid canonical HTTP(S) URL",
                        severity="warning",
                        sheet="Source Catalog",
                        row=row_number,
                    )
                )
            category = str(row.get("Primary Resource Category") or "").strip()
            if category and category not in categories:
                issues.append(
                    ValidationIssue(
                        "unknown_category",
                        f"{source_id} references unknown category: {category}",
                        severity="warning",
                        sheet="Source Catalog",
                        row=row_number,
                    )
                )
            for column in ("Authority Class", "Status", "Citation Rule", "Refresh Cadence"):
                if not str(row.get(column) or "").strip():
                    issues.append(
                        ValidationIssue(
                            "missing_governance_value",
                            f"{source_id} is missing {column}",
                            sheet="Source Catalog",
                            row=row_number,
                        )
                    )
        if guide_reported_count is not None and source_count != guide_reported_count:
            issues.append(
                ValidationIssue(
                    "guide_catalog_count_mismatch",
                    (
                        f"Workbook has {source_count} sources; guide headline reports "
                        f"{guide_reported_count}."
                    ),
                    severity="warning",
                )
            )
    finally:
        workbook.close()
    return CatalogValidationReport(workbook_path, source_count, guide_reported_count, issues)
